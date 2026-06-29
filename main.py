import os, json, re, io
from datetime import date, timedelta
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import google.generativeai as genai

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── 색상/스타일 상수 ──────────────────────────────────────────
GRAY       = PatternFill("solid", fgColor="BFBFBF")   # 헤더
GRAY_LIGHT = PatternFill("solid", fgColor="D9D9D9")   # 자료탭 헤더
YELLOW_HL  = PatternFill("solid", fgColor="FFF2CC")   # 데이터확인필요
YELLOW_HDR = PatternFill("solid", fgColor="FFFF00")   # 처리율
ORANGE     = PatternFill("solid", fgColor="FFE4B5")   # 환경셀
GREEN_OPEN = PatternFill("solid", fgColor="C6EFCE")   # OPEN
CASE_DIV   = PatternFill("solid", fgColor="D9E1F2")   # 케이스구분행

FONT_BASE   = lambda bold=False, color="000000", size=11: Font(name="맑은 고딕", size=size, bold=bold, color=color)
FONT_BLUE   = Font(name="맑은 고딕", size=11, color="0070C0")   # 오류탭 예시행
FONT_LEGEND = Font(name="맑은 고딕", size=11, color="B8860B")   # 범례 ■

ALIGN_CC  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LC  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RC  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    if num_fmt: c.number_format = num_fmt
    return c

def header_cell(ws, row, col, value, fill=None):
    set_cell(ws, row, col, value,
             font=FONT_BASE(bold=True),
             fill=fill or GRAY,
             align=ALIGN_CC,
             border=thin_border())

# ── 날짜 계산 ─────────────────────────────────────────────────
def subtract_workdays(d: date, n: int) -> date:
    while n > 0:
        d -= timedelta(days=1)
        if d.weekday() < 5:
            n -= 1
    return d

def fmt_date(d: date) -> str:
    return d.strftime("%Y-%m-%d")

# ── 탭1: 자료 및 일정 ─────────────────────────────────────────
def build_sheet1(wb, info, parsed):
    ws = wb.create_sheet("자료 및 일정")
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 45

    # 섹션 타이틀
    c = ws.cell(row=1, column=2, value="1. 자료 및 테스트 참고사항")
    c.font = FONT_BASE(bold=True)

    # 헤더
    header_cell(ws, 3, 2, "구분", fill=GRAY_LIGHT)
    header_cell(ws, 3, 3, "url",  fill=GRAY_LIGHT)
    ws.merge_cells("C3:F3")

    rows = [
        ("기획안",          "← Figma 캡처 이미지 참고"),
        ("요구사항정의서",  "← 업로드 시 입력"),
        ("테스트 데이터",   "← 업데이트 시 입력"),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        set_cell(ws, i, 2, k, font=FONT_BASE(), align=ALIGN_LC, border=thin_border())
        set_cell(ws, i, 3, v, font=FONT_BASE(), align=ALIGN_LC, border=thin_border())
        ws.merge_cells(f"C{i}:F{i}")
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 18

    # 일정 섹션
    r = 8
    c2 = ws.cell(row=r, column=2, value="2. 일정에 따른 테스트 주요 확인사항")
    c2.font = FONT_BASE(bold=True)

    r += 1
    for col, txt in [(2,"환경"),(3,"시작일"),(4,"종료일"),(5,"주요 확인사항")]:
        header_cell(ws, r, col, txt, fill=GRAY_LIGHT)

    open_d   = date.fromisoformat(info["openDate"])
    beta_end = subtract_workdays(open_d, 1)
    beta_st  = subtract_workdays(beta_end,  info["betaDays"]-1)
    alp_end  = subtract_workdays(beta_st, 1)
    alp_st   = subtract_workdays(alp_end,  info["alphaDays"]-1)
    dev_end  = subtract_workdays(alp_st, 1)
    dev_st   = subtract_workdays(dev_end,   info["devDays"]-1)

    envs = [
        ("DEV",   dev_st,  dev_end,  parsed.get("주요확인사항",{}).get("dev",  "기본 기능 검증"), ORANGE),
        ("ALPHA", alp_st,  alp_end,  parsed.get("주요확인사항",{}).get("alpha","통합 테스트"),    ORANGE),
        ("BETA",  beta_st, beta_end, parsed.get("주요확인사항",{}).get("beta", "최종 검증"),      ORANGE),
        ("OPEN",  open_d,  open_d,   "오픈 확인",                                               GREEN_OPEN),
    ]
    for i, (env, s, e, note, fill) in enumerate(envs, start=r+1):
        set_cell(ws, i, 2, env,        font=FONT_BASE(bold=True), fill=fill, align=ALIGN_CC, border=thin_border())
        set_cell(ws, i, 3, fmt_date(s), font=FONT_BASE(),          align=ALIGN_CC, border=thin_border())
        set_cell(ws, i, 4, fmt_date(e), font=FONT_BASE(),          align=ALIGN_CC, border=thin_border())
        set_cell(ws, i, 5, note,        font=FONT_BASE(),          align=ALIGN_LC, border=thin_border())

    return ws

# ── 탭2: TEST CASE ────────────────────────────────────────────
def build_sheet2(wb, info, parsed):
    ws = wb.create_sheet("TEST CASE")
    tc = parsed.get("testCases", [])
    testers = info["testers"]
    envs = ["DEV", "ALPHA", "BETA", "REAL"]

    # 열 너비
    ws.column_dimensions["A"].width = 5   # no
    ws.column_dimensions["B"].width = 18  # 구분
    ws.column_dimensions["C"].width = 18  # 영역
    ws.column_dimensions["D"].width = 14  # 회원조건
    ws.column_dimensions["E"].width = 100 # 테스트 내용
    col = 6
    for _ in envs:
        for t in testers:
            ws.column_dimensions[get_column_letter(col)].width = 9
            col += 1
            ws.column_dimensions[get_column_letter(col)].width = 9
            col += 1

    # 행1: 범례
    c = ws.cell(row=1, column=1, value="■ : 테스트 데이터 및 요구사항정의서 확인 필요")
    c.font = Font(name="맑은 고딕", size=11, color="B8860B")
    c.alignment = ALIGN_LC
    row2 = ws.cell(row=2, column=1, value="O / X / △ 결과 입력 (드롭다운)")
    row2.font = FONT_BASE()
    row2.alignment = ALIGN_LC

    # 3단 헤더 (row 3,4,5)
    fixed_cols = ["no","구분","영역","회원조건","테스트 내용"]
    for ci, v in enumerate(fixed_cols, start=1):
        for ri in [3, 4, 5]:
            header_cell(ws, ri, ci)
        ws.cell(row=3, column=ci).value = v
        ws.merge_cells(start_row=3, start_column=ci, end_row=5, end_column=ci)

    env_col = 6
    for env in envs:
        env_start = env_col
        for t in testers:
            header_cell(ws, 3, env_col, env)
            header_cell(ws, 3, env_col+1, env)
            header_cell(ws, 4, env_col,   t["name"] if len(testers)>1 else "")
            header_cell(ws, 4, env_col+1, t["name"] if len(testers)>1 else "")
            header_cell(ws, 5, env_col,   f"PC({t['pc']})")
            header_cell(ws, 5, env_col+1, f"M({t['mobile']})")
            env_col += 2
        # 환경명 병합 (row3)
        if env_col - env_start > 1:
            ws.merge_cells(start_row=3, start_column=env_start, end_row=3, end_column=env_col-1)
        # 테스터명 병합 (row4) - 테스터별 2칸씩
        for ti, t in enumerate(testers):
            sc = env_start + ti*2
            ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=sc+1)

    total_cols = env_col - 1

    # 틀고정: 5행 아래, E열 오른쪽
    ws.freeze_panes = ws.cell(row=6, column=6)

    # 드롭다운
    dv = DataValidation(type="list", formula1='"O,X,△"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.sqref = f"F6:{get_column_letter(total_cols)}10000"

    # 데이터 행
    data_row = 6
    # 병합 추적용
    merge_cols = {"B": None, "C": None, "D": None}  # {col_letter: (start_row, value)}

    def flush_merge(col_letter, current_row):
        if merge_cols[col_letter] and merge_cols[col_letter][0] < current_row - 1:
            sr = merge_cols[col_letter][0]
            col_idx = {"B":2,"C":3,"D":4}[col_letter]
            ws.merge_cells(start_row=sr, start_column=col_idx, end_row=current_row-1, end_column=col_idx)

    prev = {"구분": None, "영역": None, "회원조건": None}

    for item in tc:
        if item.get("케이스구분"):
            # 이전 병합 flush
            for col_l in ["B","C","D"]:
                flush_merge(col_l, data_row)
                merge_cols[col_l] = None
                prev = {"구분": None, "영역": None, "회원조건": None}
            # 구분행
            c = ws.cell(row=data_row, column=1, value=item.get("테스트내용",""))
            c.font = FONT_BASE(bold=True)
            c.fill = CASE_DIV
            c.alignment = ALIGN_LC
            ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=total_cols)
            ws.row_dimensions[data_row].height = 18
            data_row += 1
            continue

        fill = YELLOW_HL if item.get("데이터확인필요") else None

        # no
        set_cell(ws, data_row, 1, item.get("no",""), font=FONT_BASE(), align=ALIGN_CC, border=thin_border())

        # 구분 B
        b_val = item.get("구분","")
        set_cell(ws, data_row, 2, b_val if b_val != prev["구분"] else "",
                 font=FONT_BASE(), fill=fill, align=ALIGN_CC, border=thin_border())
        if b_val != prev["구분"]:
            flush_merge("B", data_row)
            merge_cols["B"] = (data_row, b_val)
            prev["구분"] = b_val
            prev["영역"] = None
            prev["회원조건"] = None

        # 영역 C
        c_val = item.get("영역","")
        set_cell(ws, data_row, 3, c_val if c_val != prev["영역"] else "",
                 font=FONT_BASE(), fill=fill, align=ALIGN_CC, border=thin_border())
        if c_val != prev["영역"]:
            flush_merge("C", data_row)
            merge_cols["C"] = (data_row, c_val)
            prev["영역"] = c_val
            prev["회원조건"] = None

        # 회원조건 D
        d_val = item.get("회원조건","")
        set_cell(ws, data_row, 4, d_val if d_val != prev["회원조건"] else "",
                 font=FONT_BASE(), fill=fill, align=ALIGN_CC, border=thin_border())
        if d_val != prev["회원조건"]:
            flush_merge("D", data_row)
            merge_cols["D"] = (data_row, d_val)
            prev["회원조건"] = d_val

        # 테스트 내용 E
        set_cell(ws, data_row, 5, item.get("테스트내용",""),
                 font=FONT_BASE(), fill=fill, align=ALIGN_LC, border=thin_border())

        # 환경 결과셀
        for ci in range(6, total_cols+1):
            set_cell(ws, data_row, ci, "", font=FONT_BASE(), fill=fill, align=ALIGN_CC, border=thin_border())

        ws.row_dimensions[data_row].height = None  # auto
        data_row += 1

    # 마지막 병합 flush
    for col_l in ["B","C","D"]:
        flush_merge(col_l, data_row)

    return ws

# ── 탭3: 오류 ─────────────────────────────────────────────────
def build_sheet3(wb):
    ws = wb.create_sheet("오류")
    widths = [5,8,40,15,30,10,10,10,12,12,15]
    cols   = "ABCDEFGHIJK"
    for i,w in enumerate(widths):
        ws.column_dimensions[cols[i]].width = w

    # 처리율
    set_cell(ws,1,1,"처리율", font=FONT_BASE(bold=True), fill=PatternFill("solid",fgColor="FFFF00"), align=ALIGN_CC, border=thin_border())
    set_cell(ws,1,2,"=IFERROR(COUNTA(J3:J1048576)/COUNTA(G3:G1048576),0)",
             font=FONT_BASE(bold=True), fill=PatternFill("solid",fgColor="FFFF00"),
             align=ALIGN_CC, border=thin_border(), num_fmt="0.0%")

    # 헤더
    hdrs = ["no","기기","url","영역","내용","참고이미지","테스터","처리자","처리날짜","처리결과","비고"]
    for ci, h in enumerate(hdrs, start=1):
        header_cell(ws, 2, ci, h, fill=GRAY_LIGHT)

    # 예시행 (파란색)
    example = [0,"pc","https://www-dev.catch.co.kr/event/...","기업리뷰 메인","오류 내용 기입",1,"박소연","개발자","6/4","확인완료",""]
    aligns  = [ALIGN_CC,ALIGN_CC,ALIGN_LC,ALIGN_CC,ALIGN_LC,ALIGN_CC,ALIGN_CC,ALIGN_CC,ALIGN_CC,ALIGN_CC,ALIGN_LC]
    for ci, (v, al) in enumerate(zip(example, aligns), start=1):
        set_cell(ws, 3, ci, v, font=FONT_BLUE, align=al, border=thin_border())

    # 드롭다운
    dv_device = DataValidation(type="list", formula1='"PC,M,공통"',    allow_blank=True)
    dv_result = DataValidation(type="list", formula1='"확인완료,-,추후진행"', allow_blank=True)
    ws.add_data_validation(dv_device); dv_device.sqref = "B4:B10000"
    ws.add_data_validation(dv_result); dv_result.sqref = "J4:J10000"

    ws.freeze_panes = "A3"
    return ws

# ── 탭4: 참고이미지 ───────────────────────────────────────────
def build_sheet4(wb):
    ws = wb.create_sheet("참고이미지")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30
    for ci, h in enumerate(["no","이미지","비고"], start=1):
        header_cell(ws, 1, ci)
        ws.cell(row=1, column=ci).value = h
    for r in range(2, 22):
        ws.row_dimensions[r].height = 80
        set_cell(ws, r, 1, r-1, font=FONT_BASE(), align=ALIGN_CC, border=thin_border())
        set_cell(ws, r, 2, "",   font=FONT_BASE(), align=ALIGN_CC, border=thin_border())
        set_cell(ws, r, 3, "",   font=FONT_BASE(), align=ALIGN_LC, border=thin_border())
    ws.freeze_panes = "A2"
    return ws

# ── Gemini 분석 ───────────────────────────────────────────────
def analyze_with_gemini(api_key, excel_text, images, info):
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-3.1-flash-lite")

    testers = info["testers"]
    tester_desc = "\n".join([f"{t['label']}: {t['name']} (PC: {t['pc']}, M: {t['mobile']})" for t in testers])

    prompt = f"""당신은 QA 테스트시트 전문가입니다. 아래 요구사항정의서를 분석하여 반드시 JSON만 출력하세요.

## 프로젝트 정보
- 서비스명: {info['serviceName']}
- OPEN일: {info['openDate']}
- DEV: {info['devDays']}일, ALPHA: {info['alphaDays']}일, BETA: {info['betaDays']}일
- 테스터:
{tester_desc}
- 추가 메모: {info.get('extraNote','없음')}

## 요구사항정의서
{excel_text}

## 출력 형식 (순수 JSON만)
{{
  "testCases": [
    {{
      "no": 1,
      "구분": "화면/기능 단위",
      "영역": "세부 영역",
      "회원조건": "로그인/비로그인 등",
      "테스트내용": "~되는가?\\n상세조건",
      "데이터확인필요": false,
      "케이스구분": false
    }}
  ],
  "주요확인사항": {{
    "dev": "DEV 주요 확인사항",
    "alpha": "ALPHA 주요 확인사항",
    "beta": "BETA 주요 확인사항"
  }}
}}

## 작성 규칙
1. 중복 테스트 케이스 제거
2. 글자수/자릿수 검증 제외 (기능·정책·노출 로직만)
3. 테스트내용은 반드시 "~되는가?" 형식
4. DB/정렬/노출 데이터 대조 필요한 케이스는 데이터확인필요: true
5. 비로그인 vs 로그인 등 회원조건 분기 시 케이스구분: true (앞에 구분행 삽입)
6. 케이스구분행의 테스트내용 = "Case 1) 비로그인" 같은 구분 텍스트
7. 관련 케이스는 인접 배치"""

    parts = [prompt]
    for img in images:
        import base64
        parts.append({"mime_type": img["mimeType"], "data": base64.b64decode(img["base64"])})

    response = model.generate_content(parts)
    text = response.text

    json_match = re.search(r'\{[\s\S]*\}', text)
    if not json_match:
        raise ValueError("Gemini 응답에서 JSON을 찾을 수 없습니다.")
    return json.loads(json_match.group())

# ── 메인 엔드포인트 ───────────────────────────────────────────
@app.post("/generate")
async def generate(
    file:        UploadFile = File(...),
    images:      str        = Form("[]"),
    serviceName: str        = Form(...),
    openDate:    str        = Form(...),
    devDays:     int        = Form(...),
    alphaDays:   int        = Form(...),
    betaDays:    int        = Form(...),
    testers:     str        = Form(...),
    extraNote:   str        = Form(""),
    apiKey:      str        = Form(...),
):
    # Excel 읽기
    content = await file.read()
    xl = openpyxl.load_workbook(io.BytesIO(content))
    excel_text = ""
    for sn in xl.sheetnames:
        ws = xl[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in row_vals):
                rows.append("\t".join(row_vals))
        if rows:
            excel_text += f"\n=== 시트: {sn} ===\n" + "\n".join(rows) + "\n"
    excel_text = excel_text[:14000]

    info = {
        "serviceName": serviceName,
        "openDate":    openDate,
        "devDays":     devDays,
        "alphaDays":   alphaDays,
        "betaDays":    betaDays,
        "testers":     json.loads(testers),
        "extraNote":   extraNote,
    }
    images_data = json.loads(images)

    # Gemini 분석
    parsed = analyze_with_gemini(apiKey, excel_text, images_data, info)

    # xlsx 생성
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    build_sheet1(wb, info, parsed)
    build_sheet2(wb, info, parsed)
    build_sheet3(wb)
    build_sheet4(wb)

    # 모든 시트 기본 폰트 설정
    from openpyxl.styles.named_styles import NamedStyle
    today = date.today().strftime("%y%m%d")
    filename = f"{serviceName}_테스트시트_박소연_{today}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
                 "Access-Control-Expose-Headers": "Content-Disposition"}
    )

@app.get("/")
def root():
    return {"status": "ok", "message": "QA 테스트시트 생성기 API"}
